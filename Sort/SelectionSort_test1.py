from tkinter import *
from tkinter import filedialog
from openpyxl import *
from openpyxl import Workbook
from openpyxl import load_workbook



filename = filedialog.askopenfilename(initialdir="/", title="파일선택", 
                                                filetypes=(("xlsx files","*.xlsx"),          
                                                ("all files", "*.*")))  #파일 선택

load_wb = load_workbook("%s"%filename, data_only=True) #선택한 파일 load

load_ws = load_wb['Sheet1']     #sheet선택

max_row = load_ws.max_row #열 개수

ho=[]   #리스트 선언
for i in range(max_row): #리스트에 셀값 삽입
    ho.append(load_ws.cell(i+1, 1).value)

for i in range(max_row-1):
    min=i
    for j in range(i,max_row-1):
        if ho[min]>ho[j+1]:
            min=j+1
    ho[i],ho[min]=ho[min],ho[i]
     

for i in range(max_row):     #엑셀에 입력
    load_ws.cell(i+1,2,int(ho[i]))

load_wb.save("%s"%filename)    #저장
