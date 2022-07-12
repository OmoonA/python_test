from tkinter import *
from tkinter import filedialog
from openpyxl import *
from openpyxl import Workbook
from openpyxl import load_workbook

x=0 #값의 비교를 위한 변수 선언
y=0

root = Tk()
root.title('파일 다이얼로그')

root.filename = filedialog.askopenfilename(initialdir="/", title="파일선택", 
                                                filetypes=(("xlsx files","*.xlsx"),          
                                                ("all files", "*.*")))  #파일 선택

load_wb = load_workbook("%s"%root.filename, data_only=True) #선택한 파일 load

load_ws = load_wb['Sheet1']     #sheet선택

ho=[]   #리스트 선언
for i in range(10): #리스트에 셀값 삽입
    ho.append(load_ws.cell(i+1, 1).value)
for j in range(9):  #숫자가 제자리에 배치되는 조건문 9회반복
    for i in range(0,9-j): 
        if ho[i] > ho[i+1]:#비교하여 큰숫자를 오른쪽에 두는 알고리즘
            ho[i],ho[i+1]= ho[i+1],ho[i] #신기술접목
          
print(ho) #확인용 출력

# for j in range(9):  #숫자가 제자리에 배치되는 조건문 9회반복
#     for i in range(0,9-j): 
#         if ho[i] > ho[i+1]:#비교하여 큰숫자를 오른쪽에 두는 알고리즘
#             y=ho[i] 
#             x=ho[i+1]
#         else:
#             x=ho[i] 
#             y=ho[i+1]
#         ho[i]=x 
#         ho[i+1]=y 
# print(ho) #확인용 출력

for i in range(10):     #엑셀에 입력
    load_ws.cell(i+1,2,int(ho[i]))

load_wb.save("%s"%root.filename)    #저장
