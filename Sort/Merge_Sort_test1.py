from tkinter import *
from tkinter import filedialog
from openpyxl import *
from openpyxl import Workbook
from openpyxl import load_workbook



filename = filedialog.askopenfilename(initialdir="/", title="파일선택", 
                                                filetypes=(("xlsx files","*.xlsx"),          
                                                ("all files", "*.*")))  #파일 선택

load_wb = load_workbook("%s"%filename, data_only=True) #선택한 파일 load

load_ws = load_wb['Sheet1']     #sheet선택

max_row = load_ws.max_row #열 개수

ho=[]   #리스트 선언
for i in range(max_row): #리스트에 셀값 삽입
    ho.append(load_ws.cell(i+1, 1).value)

# >>>>>>>>>
def mergesort(list):
    def sort(unsorted_list):
        if len(unsorted_list)==1:
            return
        mid=len(unsorted_list)//2
        left=unsorted_list[:mid]
        right=unsorted_list[mid:]
        mergesort(left)
        mergesort(right)
        merge(left,right)
    def merge(left,right):
        i=0
        j=0
        k=0
        while (i<len(left))and(j<len(right)):
            if left[i]<right[j]:
                list[k]=left[i]
                i+=1
                k+=1
            else:
                list[k]=right[j]
                j+=1
                k+=1
        while i<len(left):
            list[k]=left[i]
            i+=1
            k+=1
        while j<len(right):
            list[k]=right[j]
            j+=1
            k+=1
    sort(list)
    return list
mergesort(ho)
# >>>>>>>>>>>

for i in range(max_row):     #엑셀에 입력
    load_ws.cell(i+1,2,int(ho[i]))

load_wb.save("%s"%filename)    #저장
