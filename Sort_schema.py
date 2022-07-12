from tkinter import *
from tkinter import filedialog
from openpyxl import *
from openpyxl import Workbook
from openpyxl import load_workbook

x=0 #값의 비교를 위한 변수 선언
y=0

root = Tk()
root.title('파일 다이얼로그')

root.filename = filedialog.askopenfilename(initialdir="/", title="파일선택", 
                                                filetypes=(("xlsx files","*.xlsx"),          
                                                ("all files", "*.*")))  #파일 선택

load_wb = load_workbook("%s"%root.filename, data_only=True) #선택한 파일 load

load_ws = load_wb['Sheet1']     #sheet선택

ho=[]   #리스트 선언
for i in range(10): #리스트에 셀값 삽입
    ho.append(load_ws.cell(i+1, 1).value)

# >>>>>>


# >>>>>>

for i in range(10):     #엑셀에 입력
    load_ws.cell(i+1,2,int(ho[i]))

load_wb.save("%s"%root.filename)    #저장